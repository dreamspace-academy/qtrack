#Import Modules
import cv2
import openpyxl
from pyzbar import pyzbar
from datetime import datetime


#Initializing
prev_barcode_data = ""
headers = ["No", "Date", "Staff Id", "In-Time", "Out-Time"]
staff_data_path = "0-data/qTrack-staffs.xlsx"
attendance_path = "1-attendance/qTrack-attendance.xlsx"

#Read the Staff data excel
def read_excel(path, column):
    staff_ids = []
    
    staff_data_wb = openpyxl.load_workbook(path)

    staff_data_sheet = staff_data_wb.active
    max_row = staff_data_sheet.max_row


    for i in range(2, max_row + 1):
        cell = staff_data_sheet.cell(row = i, column = column)
        staff_ids.append(str(cell.value))

    return staff_ids

# Get current date and time
def get_current_time_data():
    current_date_and_time = datetime.now()
    current_time = current_date_and_time.strftime("%H:%M:%S")
    current_date = current_date_and_time.strftime("%Y-%m-%d")
    
    return [current_time, current_date]


active_staff_ids = read_excel(staff_data_path, 2)
existing_staff_ids = read_excel(attendance_path, 3)
existing_counts = read_excel(attendance_path, 1)
max_num = int(max(existing_counts)) + 1
existing_dates = read_excel(attendance_path, 2)
max_date = max(existing_dates)


#Initializing the attendance excel sheet
attendance_wb = openpyxl.load_workbook(attendance_path)
attendance_sheet = attendance_wb.active

for index, header in enumerate(headers):
    cell = attendance_sheet.cell(row = 1, column = index+1)
    cell.value = header

cap = cv2.VideoCapture(0)

while True:
    ret, frame = cap.read()
    barcodes = pyzbar.decode(frame)

    for barcode in barcodes:
        (x, y, w, h) = barcode.rect
        cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
        barcodeData = barcode.data.decode("utf-8")
        barcodeType = barcode.type
        staff_id = str(barcode.data).split("'")[1]
        cv2.putText(frame, staff_id, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)

        if barcodeData != prev_barcode_data:
            current_time, current_date = get_current_time_data() 
            prev_barcode_data = barcodeData

            print(staff_id)

            if staff_id in active_staff_ids:

                if staff_id in existing_staff_ids :
                    for index, id in enumerate(existing_staff_ids):
                        cell_staff = attendance_sheet.cell(row = index+2, column = 3)
                        cell_date = attendance_sheet.cell(row = index+2, column = 2)
                        if (cell_staff.value == staff_id) and (cell_date.value == max_date):
                            cell_a = attendance_sheet.cell(row = index+2, column = 5)
                            cell_a.value = current_time

                else:
                    attendance_sheet.insert_rows(idx=2)

                    cell_a1 = attendance_sheet.cell(row = 2, column = 1)
                    cell_a1.value = max_num
                    max_num += 1

                    data = [current_date, staff_id, current_time]

                    for i in range(3):
                        cell_a2 = attendance_sheet.cell(row = 2, column = i+2)
                        cell_a2.value = data[i]

                attendance_wb.save(attendance_path)

    cv2.imshow("Webcam", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break
    
cap.release()
cv2.destroyAllWindows()